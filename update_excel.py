from openpyxl import load_workbook
from datetime import date,timedelta
import os
import openpyxl
import update_chart
# Returns the current local date

def check(email_from,email_to,password,emlcheck):
    today = date.today()
    PATH='Excel/'+str(today)+'.xlsx';
    PrevPATH='Excel/'+str(today-timedelta(days=1))+'.xlsx';
    if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
        book=load_workbook(PATH)
        sheet=book.active
        return book,sheet
        
    else:
        book=openpyxl.Workbook();
        sheet=book.active
        sheet['A1']="Time";
        sheet['B1']="Animal";
        sheet['C1']="Numbers";
        book.save(PATH);
        if(os.path.isfile(PrevPATH) and os.access(PrevPATH, os.R_OK)):
            update_chart.uc(email_from,email_to,password,emlcheck,PrevPATH);
        return book,sheet

#print("Today date is: ", today)

def update_ex(email_from,email_to,password,emlcheck,data,ct):
    
    
    today = date.today()
    book,sheet=check(email_from,email_to,password,emlcheck);
    new_data=data;
    r=sheet.max_row+1;
    count=0;
    for i in range(r,r+len(new_data[0])):
        
        sheet['A'+str(r+count)]=ct;
        sheet['B'+str(r+count)]=new_data[0][count];
        sheet['c'+str(r+count)]=new_data[1][count];
        count+=1;
    book.save('Excel/'+str(today)+'.xlsx')
    
